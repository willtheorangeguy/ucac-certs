from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import theme
from .awards import COLUMNS
from .grid import EXPIRY_WARNING_DAYS, Grid, MemberRow
from .models import CellStatus

DATE_FORMAT = "yyyy-mm-dd"
HEADINGS = ["Names:", *(column.code for column in COLUMNS), "LS#"]
_THIN = Side(style="thin", color="AEAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _write_preamble(sheet: Worksheet, grid: Grid) -> int:
    sheet["A1"] = "Ucalgary Aquatic Center Staff Certifications"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Generated {grid.as_of.isoformat()}"

    row = 4
    for column in COLUMNS:
        years = column.validity_years
        sheet.cell(row=row, column=1, value=f"{column.code} - {column.label}")
        sheet.cell(
            row=row,
            column=2,
            value=f"Current for {years} year{'s' if years != 1 else ''} from certification date",
        )
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Award expired").fill = _fill(theme.EXPIRED)
    sheet.cell(row=row, column=2, value=f"Expires within {EXPIRY_WARNING_DAYS} days").fill = _fill(
        theme.EXPIRING
    )
    missing = sheet.cell(row=row, column=3, value="No award on record")
    missing.fill = _fill(theme.MISSING)
    missing.font = Font(color=theme.STATUS_TEXT[CellStatus.MISSING])
    return row + 2


def _write_row(sheet: Worksheet, row_index: int, row: MemberRow) -> None:
    name = sheet.cell(row=row_index, column=1, value=row.name)
    name.border = _BORDER
    if row.name_warning:
        name.font = Font(italic=True)

    for offset, cell in enumerate(row.cells, start=2):
        target = sheet.cell(row=row_index, column=offset)
        if cell.expiry_date is not None:
            target.value = cell.expiry_date
            target.number_format = DATE_FORMAT
        fill = theme.STATUS_FILL[cell.status]
        if fill:
            target.fill = _fill(fill)
        target.font = Font(
            color=theme.STATUS_TEXT[cell.status], italic=cell.provisional
        )
        target.alignment = Alignment(horizontal="center")
        target.border = _BORDER

    code = sheet.cell(row=row_index, column=len(HEADINGS), value=row.member_code)
    code.alignment = Alignment(horizontal="center")
    code.border = _BORDER

    if row.error:
        for column_index in range(2, len(HEADINGS)):
            sheet.cell(row=row_index, column=column_index).fill = _fill(theme.ERROR)
        sheet.cell(row=row_index, column=2).value = row.error


def _write_diagnostics(sheet: Worksheet, grid: Grid) -> None:
    sheet.append(["Type", "Detail"])
    sheet["A1"].font = sheet["B1"].font = Font(bold=True)
    for row in grid.rows:
        if row.error:
            sheet.append(["Lookup error", f"{row.name} ({row.member_code}): {row.error}"])
        if row.name_warning:
            sheet.append(["Name warning", f"{row.member_code}: {row.name_warning}"])
    for title in grid.unmapped_awards:
        sheet.append(["Unmapped award", title])
    for note in grid.disagreements:
        sheet.append(["Status disagreement", note])
    for row in grid.rows:
        for cell in row.cells:
            if cell.provisional:
                sheet.append(
                    [
                        "Provisional",
                        f"{row.name} {cell.column.code} derived from {cell.source_award}",
                    ]
                )
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 90


def build_workbook(grid: Grid, output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certifications"

    header_row = _write_preamble(sheet, grid)
    for index, heading in enumerate(HEADINGS, start=1):
        cell = sheet.cell(row=header_row, column=index, value=heading)
        cell.font = Font(bold=True)
        cell.fill = _fill(theme.HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    row_index = header_row + 1
    seen_away = False
    for row in grid.rows:
        if row.away and not seen_away:
            seen_away = True
            marker = sheet.cell(row=row_index, column=1, value="Away Spring / Summer")
            marker.font = Font(bold=True)
            marker.fill = _fill(theme.HEADER)
            row_index += 1
        _write_row(sheet, row_index, row)
        row_index += 1

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=2)
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADINGS))}{row_index - 1}"
    sheet.column_dimensions["A"].width = 26
    for index in range(2, len(HEADINGS) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 13

    _write_diagnostics(workbook.create_sheet("Diagnostics"), grid)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
