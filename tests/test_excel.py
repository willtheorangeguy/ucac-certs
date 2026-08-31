from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from lss_report import theme
from lss_report.awards import COLUMNS, NATIONAL_LIFEGUARD, OXYGEN, SWIM_INSTRUCTOR
from lss_report.excel import HEADINGS, build_workbook
from lss_report.grid import build_grid
from lss_report.models import Certification, MemberRecord, ReportData

AS_OF = datetime(2026, 8, 30, 6, 15)


def _workbook(tmp_path: Path, *records: MemberRecord):
    grid = build_grid(ReportData(generated_at=AS_OF, records=list(records)))
    output = tmp_path / "report.xlsx"
    build_workbook(grid, output)
    return load_workbook(output)


def _header_row(sheet) -> int:
    for row in sheet.iter_rows(min_col=1, max_col=1):
        if row[0].value == "Names:":
            return row[0].row
    raise AssertionError("header row not found")


def test_statuses_map_to_the_reference_sheet_colours(tmp_path: Path):
    record = MemberRecord(
        configured_name="Example Staff Member",
        member_code="ABC123",
        certifications=[
            # Expired, expiring within 30 days, and no O2 award at all.
            Certification("National Lifeguard - Pool", date(2022, 1, 1), NATIONAL_LIFEGUARD, date(2024, 1, 1), site_current=False),
            Certification("Swim Instructor", date(2024, 9, 10), SWIM_INSTRUCTOR, date(2026, 9, 10)),
        ],
    )
    sheet = _workbook(tmp_path, record)["Certifications"]
    row = _header_row(sheet) + 1
    codes = [column.code for column in COLUMNS]

    def fill(code: str) -> str:
        return sheet.cell(row=row, column=2 + codes.index(code)).fill.fgColor.rgb[-6:]

    assert fill("NL") == theme.EXPIRED
    assert fill("SI") == theme.EXPIRING
    assert fill("O2") == theme.MISSING


def test_dates_are_written_as_real_dates_in_iso_format(tmp_path: Path):
    record = MemberRecord(
        configured_name="Example Staff Member",
        member_code="ABC123",
        certifications=[
            Certification("O2 Administration", date(2025, 9, 7), OXYGEN, date(2027, 9, 7))
        ],
    )
    sheet = _workbook(tmp_path, record)["Certifications"]
    cell = sheet.cell(row=_header_row(sheet) + 1, column=2 + [c.code for c in COLUMNS].index("O2"))
    assert cell.value == datetime(2027, 9, 7)
    assert cell.number_format == "yyyy-mm-dd"


def test_ls_number_is_the_last_column(tmp_path: Path):
    record = MemberRecord(configured_name="Example Staff Member", member_code="ABC123")
    sheet = _workbook(tmp_path, record)["Certifications"]
    header = _header_row(sheet)
    assert sheet.cell(row=header, column=len(HEADINGS)).value == "LS#"
    assert sheet.cell(row=header + 1, column=len(HEADINGS)).value == "ABC123"


def test_away_staff_get_a_section_marker(tmp_path: Path):
    sheet = _workbook(
        tmp_path,
        MemberRecord(configured_name="Zoe Active", member_code="AAA111"),
        MemberRecord(configured_name="Aaron Away", member_code="BBB222", away=True),
    )["Certifications"]
    header = _header_row(sheet)
    assert sheet.cell(row=header + 1, column=1).value == "Zoe Active"
    assert sheet.cell(row=header + 2, column=1).value == "Away Spring / Summer"
    assert sheet.cell(row=header + 3, column=1).value == "Aaron Away"


def test_diagnostics_sheet_lists_unmapped_awards(tmp_path: Path):
    record = MemberRecord(
        configured_name="Example Staff Member",
        member_code="ABC123",
        certifications=[Certification("Wilderness Guide Level 4", date(2025, 1, 1))],
    )
    sheet = _workbook(tmp_path, record)["Diagnostics"]
    rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows()]
    assert ("Unmapped award", "Wilderness Guide Level 4") in rows
